"""Generatore xlsx mensile AGHC — 18 tab (uno per cliente) con KPI + rational + % budget.

Aggiornamento 25/04/2026: roster esteso a 18 clienti (aggiunto Montemagno).
Lista clienti caricata da anagrafica.CLIENTS — basta aggiornare quel file per propagare ovunque."""
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from anagrafica import CLIENTS, BUDGET_WEIGHTS, META_SHARE, TIKTOK_SHARE, MONTH_IT, comparison_period

# Soglie operative
TIKTOK_MIN_MONTHLY = 600.0  # € — ogni cliente con TikTok attivo spende almeno 600 €/mese

# ---------- Stili ----------
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
KPI_TITLE_FONT = Font(bold=True, size=11, color="1F4E78")
TABLE_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TABLE_HEADER_FONT = Font(bold=True, size=10)
POSITIVE_FONT = Font(color="548235", bold=True)   # verde scuro
NEGATIVE_FONT = Font(color="C00000", bold=True)   # rosso
NEUTRAL_FONT = Font(color="595959")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------- Helpers ----------
def fmt_int(x):
    if x is None:
        return "n/d"
    return f"{int(round(x)):,}".replace(",", ".")


def fmt_eur(x):
    if x is None:
        return "n/d"
    return f"{x:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")


def pct(cur, prev):
    if prev is None or prev == 0 or cur is None:
        return None
    return (cur - prev) / prev * 100


def fmt_pct(p):
    if p is None:
        return "n/d"
    sign = "+" if p > 0 else ""
    return f"{sign}{p:.2f}%"


def pct_font(p):
    if p is None:
        return NEUTRAL_FONT
    return POSITIVE_FONT if p >= 0 else NEGATIVE_FONT


def write_kpi_table(ws, start_row, title, rows, period_a_label, period_b_label, value_fmt):
    """Scrive una tabella KPI a partire da start_row.
    rows = [(label, cur, prev), ...]
    Ritorna prossima riga disponibile."""
    ws.cell(row=start_row, column=1, value=title).font = KPI_TITLE_FONT
    header_row = start_row + 1
    headers = ["", period_a_label, period_b_label, "Confronto"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = TABLE_HEADER_FONT
        c.fill = TABLE_HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    for i, (label, cur, prev) in enumerate(rows):
        r = header_row + 1 + i
        c1 = ws.cell(row=r, column=1, value=label); c1.font = TABLE_HEADER_FONT; c1.alignment = LEFT; c1.border = BORDER
        c2 = ws.cell(row=r, column=2, value=value_fmt(cur)); c2.alignment = CENTER; c2.border = BORDER
        c3 = ws.cell(row=r, column=3, value=value_fmt(prev)); c3.alignment = CENTER; c3.border = BORDER
        p = pct(cur, prev)
        c4 = ws.cell(row=r, column=4, value=fmt_pct(p)); c4.alignment = CENTER; c4.border = BORDER; c4.font = pct_font(p)
    return header_row + 1 + len(rows) + 1


def _art_pct(n):
    """'il X%' o 'l'X%' a seconda della pronuncia italiana del numero."""
    v = int(round(abs(n)))
    if v in (1, 8, 11) or (80 <= v <= 89):
        return f"l'{v}%"
    return f"il {v}%"


def _del_pct(n):
    """'del X%' o 'dell'X%'."""
    v = int(round(abs(n)))
    if v in (1, 8, 11) or (80 <= v <= 89):
        return f"dell'{v}%"
    return f"del {v}%"


def _cross_channel_insight(p1_kind, fb_reach, fb_reach_delta, ig_reach, ig_reach_delta):
    """Se p1 ha raccontato un canale, restituisce una frase sintetica sull'altro.
    Tono sempre strategico — i cali sono inquadrati come allocazione di budget."""
    is_fb_lead = p1_kind in ("fb_explode", "fb_modest")
    is_ig_lead = p1_kind in ("ig_explode", "ig_modest")
    if is_fb_lead and ig_reach >= 5000:
        if ig_reach_delta is not None and ig_reach_delta < -25:
            return (f" Instagram resta presidiato in modo mirato ({fmt_int(ig_reach)} utenti unici raggiunti), "
                    f"con il budget concentrato su Facebook dove sta rendendo di più nel mese.")
        if ig_reach_delta is not None and ig_reach_delta > 20:
            return (f" Instagram aggiunge in parallelo {fmt_int(ig_reach)} utenti unici raggiunti "
                    f"(+{ig_reach_delta:.0f}%), confermando una presenza qualificata sul pubblico mobile-first.")
        return (f" Instagram mantiene un presidio complementare di {fmt_int(ig_reach)} utenti unici, "
                f"a copertura di un pubblico differenziato per età e abitudini di consumo.")
    if is_ig_lead and fb_reach >= 5000:
        if fb_reach_delta is not None and fb_reach_delta < -25:
            return (f" Facebook resta presidiato in modo mirato ({fmt_int(fb_reach)} utenti unici raggiunti), "
                    f"con il budget concentrato su Instagram dove sta rendendo di più nel mese.")
        if fb_reach_delta is not None and fb_reach_delta > 20:
            return (f" Facebook aggiunge in parallelo {fmt_int(fb_reach)} utenti unici raggiunti "
                    f"(+{fb_reach_delta:.0f}%), a conferma di una presenza multicanale coerente.")
        return (f" Facebook mantiene un presidio complementare di {fmt_int(fb_reach)} utenti unici, "
                f"garantendo copertura sul pubblico più ampio della piattaforma.")
    return ""


def build_rational(client_name, data_meta_cur, data_meta_prev, data_tk_cur, data_tk_prev,
                   period_a_label, period_b_label, confronto_meta, spent_total, expected_total, month_num, has_tiktok):
    """Rational 3-paragrafi, TOV professionale e polarizzante.
    Struttura: cos'è successo · perché conta · cosa stiamo facendo dopo."""
    month_low = MONTH_IT[month_num].lower()
    month_cap = MONTH_IT[month_num]
    next_month = (month_num % 12) + 1
    next_month_low = MONTH_IT[next_month].lower()
    next_month_cap = MONTH_IT[next_month]
    next_weight = BUDGET_WEIGHTS[next_month]

    # Segnali
    fb_reach_cur = data_meta_cur["facebook"]["reach"] or 0
    fb_reach_prev = data_meta_prev["facebook"]["reach"] or 0
    fb_reach_delta = pct(fb_reach_cur, fb_reach_prev)
    ig_reach_cur = data_meta_cur["instagram"]["reach"] or 0
    ig_reach_prev = data_meta_prev["instagram"]["reach"] or 0
    ig_reach_delta = pct(ig_reach_cur, ig_reach_prev)
    reach_cur_meta = fb_reach_cur + ig_reach_cur
    reach_prev_meta = fb_reach_prev + ig_reach_prev
    reach_meta_delta = pct(reach_cur_meta, reach_prev_meta)

    fb_eng_cur = data_meta_cur["facebook"].get("actions_page_engagement") or 0
    fb_eng_prev = data_meta_prev["facebook"].get("actions_page_engagement") or 0
    fb_eng_delta = pct(fb_eng_cur, fb_eng_prev)
    ig_eng_cur = data_meta_cur["instagram"].get("actions_page_engagement") or 0
    ig_eng_prev = data_meta_prev["instagram"].get("actions_page_engagement") or 0
    ig_eng_delta = pct(ig_eng_cur, ig_eng_prev)
    tot_eng = fb_eng_cur + ig_eng_cur

    spend_cur = (data_meta_cur["facebook"]["spend"] or 0) + (data_meta_cur["instagram"]["spend"] or 0)
    spend_prev = (data_meta_prev["facebook"]["spend"] or 0) + (data_meta_prev["instagram"]["spend"] or 0)
    spend_delta = pct(spend_cur, spend_prev)

    tk_launched = (
        has_tiktok and data_tk_cur and (data_tk_cur.get("spend") or 0) > 0
        and (not data_tk_prev or ((data_tk_prev.get("spend") or 0) == 0 and (data_tk_prev.get("impressions") or 0) == 0))
    )
    tk_impr_cur = (data_tk_cur.get("impressions") if has_tiktok and data_tk_cur else 0) or 0
    tk_reach_cur = (data_tk_cur.get("reach") if has_tiktok and data_tk_cur else 0) or 0

    # Caso speciale: account a zero per scelta strategica
    if spend_cur == 0 and reach_cur_meta == 0:
        p1 = f"{month_cap} rappresenta una pausa strategica per {client_name}, coerente con il calendario annuo del piano media."
        p2 = "Il budget residuo resta integro e pronto a concentrarsi sulle finestre stagionali a più alto ritorno previste nei mesi successivi."
        if next_weight >= 12:
            p3 = f"Da {next_month_low} (peso {next_weight}% del piano annuo) il presidio riprende nel cuore della stagione, con la pressione necessaria a intercettare la domanda attiva di prenotazione."
        else:
            p3 = f"Da {next_month_low} (peso {next_weight}% del piano annuo) il presidio riprende in modo progressivo, pronto a salire sulle finestre più strategiche."
        return f"{p1}\n\n{p2}\n\n{p3}"

    # Paragrafo 1
    p1_kind = None  # tag: tk_launch | fb_explode | ig_explode | reach_big | efficiency | fb_modest | ig_modest | tk_present | neutral
    if tk_launched and tk_impr_cur > 0:
        p1 = (f"{month_cap} inaugura una nuova fase per {client_name}: la prima campagna TikTok va live e debutta con "
              f"{fmt_int(tk_impr_cur)} visualizzazioni e {fmt_int(tk_reach_cur)} utenti unici raggiunti, aprendo un canale "
              f"fino a ieri inesplorato dal brand.")
        p1_kind = "tk_launch"
    elif fb_reach_delta is not None and fb_reach_delta > 100 and (reach_meta_delta is None or reach_meta_delta > 0):
        mul = f"{(1 + fb_reach_delta/100):.1f}".replace('.', ',')
        p1 = (f"{month_cap} segna un cambio di passo per {client_name}: Facebook diventa il canale di traino e amplia la "
              f"copertura di {mul} volte rispetto al periodo di confronto, raggiungendo {fmt_int(fb_reach_cur)} utenti unici.")
        p1_kind = "fb_explode"
    elif ig_reach_delta is not None and ig_reach_delta > 80 and ig_reach_cur > 50000:
        mul = f"{(1 + ig_reach_delta/100):.1f}".replace('.', ',')
        p1 = (f"{month_cap} è il mese di Instagram per {client_name}: la copertura cresce di {mul} volte, portando il "
              f"messaggio del brand davanti a {fmt_int(ig_reach_cur)} utenti unici sul canale.")
        p1_kind = "ig_explode"
    elif reach_meta_delta is not None and reach_meta_delta > 30:
        p1 = (f"{month_cap} è il mese in cui la copertura Meta cresce in modo netto: il brand entra nello sguardo di "
              f"{fmt_int(reach_cur_meta)} persone (+{reach_meta_delta:.0f}% rispetto al periodo di confronto).")
        p1_kind = "reach_big"
    elif spend_delta is not None and spend_delta < -8 and reach_meta_delta is not None and reach_meta_delta > -8:
        p1 = (f"{month_cap} è il mese dell'efficienza per {client_name}: la copertura Meta resta solida con "
              f"{fmt_int(reach_cur_meta)} utenti unici raggiunti, con un investimento ridotto {_del_pct(spend_delta)}. "
              f"Ogni euro speso ha lavorato meglio.")
        p1_kind = "efficiency"
    elif fb_reach_delta is not None and fb_reach_delta > 30:
        p1 = (f"{month_cap} consolida la presenza di {client_name} su Facebook: {fmt_int(fb_reach_cur)} utenti unici "
              f"raggiunti, +{fb_reach_delta:.0f}% rispetto al periodo di confronto.")
        p1_kind = "fb_modest"
    elif ig_reach_delta is not None and ig_reach_delta > 30:
        p1 = (f"{month_cap} rafforza la presenza di {client_name} su Instagram: {fmt_int(ig_reach_cur)} utenti unici "
              f"raggiunti, +{ig_reach_delta:.0f}% rispetto al periodo di confronto.")
        p1_kind = "ig_modest"
    elif has_tiktok and tk_impr_cur > 100000:
        p1 = (f"{month_cap} vede {client_name} attivo su entrambi i canali: Meta porta il brand davanti a "
              f"{fmt_int(reach_cur_meta)} utenti unici, TikTok aggiunge {fmt_int(tk_impr_cur)} visualizzazioni a "
              f"presidio del pubblico più giovane.")
        p1_kind = "tk_present"
    else:
        p1 = (f"{month_cap} mantiene il presidio di {client_name} su base solida: {fmt_int(reach_cur_meta)} utenti unici "
              f"raggiunti su Meta, in linea con il posizionamento strategico del mese all'interno del piano annuo.")
        p1_kind = "neutral"

    # Paragrafo 2 — evita di ripetere lo spend delta se p1 lo ha già menzionato
    spend_mentioned_in_p1 = (p1_kind == "efficiency")

    p2_parts = []
    if fb_eng_delta is not None and fb_eng_delta > 100 and fb_eng_cur > 5000:
        p2_parts.append(f"le interazioni sui contenuti Facebook salgono a {fmt_int(fb_eng_cur)}, una scala completamente "
                        f"diversa rispetto al periodo di confronto")
    elif ig_eng_delta is not None and ig_eng_delta > 40 and ig_eng_cur > 2000:
        p2_parts.append(f"le interazioni Instagram crescono {_del_pct(ig_eng_delta)} e portano il coinvolgimento totale "
                        f"a {fmt_int(ig_eng_cur)} azioni nel mese")
    elif tot_eng > 50000:
        p2_parts.append(f"il pubblico interagisce attivamente con i contenuti del brand, con {fmt_int(tot_eng)} interazioni "
                        f"totali generate nel mese")
    if has_tiktok and not tk_launched and tk_impr_cur > 100000:
        p2_parts.append(f"TikTok continua a presidiare con efficienza il pubblico più giovane "
                        f"({fmt_int(tk_impr_cur)} visualizzazioni)")

    if p2_parts:
        sent = "; ".join(p2_parts[:2])
        p2 = sent[0].upper() + sent[1:] + "."
        if not spend_mentioned_in_p1:
            if spend_delta is not None and spend_delta < -8:
                p2 += f" Risultati ottenuti con {_art_pct(spend_delta)} di investimento in meno rispetto al periodo di confronto."
            elif spend_delta is not None and spend_delta > 12:
                p2 += f" Per sostenere questa traiettoria, l'investimento del mese cresce {_del_pct(spend_delta)}."
    elif spend_delta is not None and spend_delta < -8 and not spend_mentioned_in_p1:
        p2 = (f"Il dato chiave del mese è l'efficienza: la copertura resta in linea con il periodo di confronto, ma "
              f"con {_art_pct(spend_delta)} di budget in meno. È il segnale di una strategia di targeting che continua "
              f"a lavorare bene.")
    elif spend_delta is not None and spend_delta > 12:
        p2 = (f"L'investimento del mese sale {_del_pct(spend_delta)} per consolidare il momentum, coerente con il peso "
              f"di {BUDGET_WEIGHTS[month_num]}% che il piano annuo riserva a {month_low}.")
    elif tot_eng > 0:
        p2 = (f"Il presidio del marchio si mantiene attivo con {fmt_int(tot_eng)} interazioni totali su Meta: i contenuti "
              f"pubblicati continuano a generare conversazioni qualificate intorno a {client_name}.")
    else:
        p2 = ("Il piano del mese resta coerente con la stagionalità: presidio costante a tutela della brand awareness, "
              "in attesa delle finestre più strategiche dei mesi successivi.")

    # Cross-channel insight: aggiunge una frase sull'altro canale se p1 era mono-canale
    p2 += _cross_channel_insight(p1_kind, fb_reach_cur, fb_reach_delta, ig_reach_cur, ig_reach_delta)

    # Paragrafo 3
    if tk_launched:
        nxt_phrase = "centrale per costruire la pressione stagionale" if next_weight >= 12 else "utile a consolidare il sistema appena avviato"
        p3 = (f"Da {next_month_low} {client_name} opera su due leve complementari — Meta per la conversione, TikTok per la "
              f"scoperta. {next_month_cap} pesa il {next_weight}% del piano annuo, una finestra {nxt_phrase}.")
    elif has_tiktok:
        p3 = (f"A {next_month_low} (peso {next_weight}% del piano annuo) confermiamo il sistema a due velocità: Meta "
              f"presidia la fase di considerazione e prenotazione, TikTok amplia la scoperta del brand sul pubblico più giovane.")
    elif next_weight >= 12:
        p3 = (f"{next_month_cap} entra nel cuore della stagione ({next_weight}% del budget annuo): saliamo sulla pressione "
              f"per intercettare la domanda attiva nella finestra prenotativa più calda dell'anno.")
    elif next_weight >= 10:
        p3 = (f"A {next_month_low} ({next_weight}% del piano annuo) la pressione cresce in modo strutturato: prepariamo "
              f"il pubblico alle finestre di alta stagione che arrivano subito dopo.")
    else:
        p3 = (f"A {next_month_low} (peso {next_weight}% del piano annuo) il presidio prosegue strategico, mantenendo il "
              f"pubblico caldo in vista delle finestre più rilevanti del piano.")

    return f"{p1}\n\n{p2}\n\n{p3}"


def build_client_sheet(wb, client, data, year, month):
    """Costruisce una scheda per un singolo cliente."""
    sheet_name = client["nome"][:31]  # Excel max 31 chars
    ws = wb.create_sheet(title=sheet_name)

    period_a_label = f"{MONTH_IT[month]} {year}"
    comp_y, comp_m = comparison_period(year, month, client["confronto_meta"])
    period_b_label = f"{MONTH_IT[comp_m]} {comp_y}"

    # Recupero dati Meta
    if client["meta_filter"]:
        # cliente su account condiviso
        fdata = data["meta_filtered"].get(client["nome"], {})
        meta_cur = fdata.get("current", {"facebook": {}, "instagram": {}})
        key_prev = "prev_yoy" if client["confronto_meta"] == "YoY" else "prev_mom"
        meta_prev = fdata.get(key_prev, {"facebook": {}, "instagram": {}})
    else:
        by_account = data["meta_by_account"]
        acc_id = client["meta_account_id"]
        meta_cur = by_account["current"].get(acc_id, {"facebook": {}, "instagram": {}})
        key_prev = "prev_yoy" if client["confronto_meta"] == "YoY" else "prev_mom"
        meta_prev = by_account[key_prev].get(acc_id, {"facebook": {}, "instagram": {}})

    # Normalizza dict vuoti
    def ensure(d, plat):
        k = d.get(plat, {})
        if not k:
            return {"reach": 0, "impressions": 0, "actions_page_engagement": 0, "clicks": 0, "spend": 0}
        return k
    meta_cur = {"facebook": ensure(meta_cur, "facebook"), "instagram": ensure(meta_cur, "instagram")}
    meta_prev = {"facebook": ensure(meta_prev, "facebook"), "instagram": ensure(meta_prev, "instagram")}

    # Recupero dati TikTok
    has_tiktok = bool(client["tiktok_account_id"])
    tk_cur = tk_prev = None
    if has_tiktok:
        tk_id = client["tiktok_account_id"]
        tk_cur_raw = data["tiktok"]["current"].get(tk_id)
        tk_prev_raw = data["tiktok"].get("prev_mom", {}).get(tk_id) or data["tiktok"].get("prev_yoy", {}).get(tk_id)
        tk_cur = tk_cur_raw if tk_cur_raw else {"reach": 0, "impressions": 0, "engagements": 0, "clicks": 0, "spend": 0}
        tk_prev = tk_prev_raw if tk_prev_raw else {"reach": 0, "impressions": 0, "engagements": 0, "clicks": 0, "spend": 0}

    # ---------- Scrittura sheet ----------
    # Titolo
    ws.cell(row=1, column=1, value=f"{client['nome']} — {period_a_label} vs {period_b_label}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # Riga anagrafica/budget cliente (sempre)
    next_row = 2
    tk_status = "attivo" if has_tiktok else ("non gestito")
    info_line = (f"Budget annuo: {fmt_eur(client['budget_annuo'])}  ·  Confronto Meta: {client['confronto_meta']}"
                 f"  ·  TikTok: {tk_status}"
                 + (f"  ·  Confronto TikTok: {client['confronto_tiktok']}" if client['confronto_tiktok'] else ""))
    ws.cell(row=next_row, column=1, value=info_line).font = Font(italic=True, color="1F4E78", size=10, bold=True)
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
    next_row += 1
    if client["note"]:
        ws.cell(row=next_row, column=1, value=f"📌 {client['note']}").font = Font(italic=True, color="595959", size=9)
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
        next_row += 1

    # Sezione META
    ws.cell(row=next_row + 1, column=1, value="═══ META (Facebook + Instagram) ═══").font = Font(bold=True, size=12, color="1F4E78")
    ws.merge_cells(start_row=next_row + 1, start_column=1, end_row=next_row + 1, end_column=4)
    next_row += 3

    for kpi_title, field, fmt in [
        ("ACCOUNT RAGGIUNTI", "reach", fmt_int),
        ("VISUALIZZAZIONI", "impressions", fmt_int),
        ("INTERAZIONI", "actions_page_engagement", fmt_int),
        ("CLICKS", "clicks", fmt_int),
    ]:
        next_row = write_kpi_table(
            ws, next_row, kpi_title,
            [("Instagram", meta_cur["instagram"].get(field, 0), meta_prev["instagram"].get(field, 0)),
             ("Facebook", meta_cur["facebook"].get(field, 0), meta_prev["facebook"].get(field, 0))],
            period_a_label, period_b_label, fmt,
        )

    # Budget Meta (totale)
    spend_cur = (meta_cur["facebook"]["spend"] or 0) + (meta_cur["instagram"]["spend"] or 0)
    spend_prev = (meta_prev["facebook"]["spend"] or 0) + (meta_prev["instagram"]["spend"] or 0)
    next_row = write_kpi_table(
        ws, next_row, "BUDGET META",
        [("Totale", spend_cur, spend_prev)],
        period_a_label, period_b_label, fmt_eur,
    )

    # Sezione TikTok
    if has_tiktok:
        ws.cell(row=next_row + 1, column=1, value="═══ TIKTOK ═══").font = Font(bold=True, size=12, color="1F4E78")
        ws.merge_cells(start_row=next_row + 1, start_column=1, end_row=next_row + 1, end_column=4)
        next_row += 3

        comp_y_t, comp_m_t = comparison_period(year, month, client["confronto_tiktok"])
        period_b_tk = f"{MONTH_IT[comp_m_t]} {comp_y_t}"

        for kpi_title, field, fmt in [
            ("ACCOUNT RAGGIUNTI", "reach", fmt_int),
            ("VISUALIZZAZIONI", "impressions", fmt_int),
            ("INTERAZIONI", "engagements", fmt_int),
            ("CLICKS", "clicks", fmt_int),
            ("BUDGET TIKTOK", "spend", fmt_eur),
        ]:
            next_row = write_kpi_table(
                ws, next_row, kpi_title,
                [("TikTok", tk_cur.get(field, 0) if tk_cur else 0, tk_prev.get(field, 0) if tk_prev else 0)],
                period_a_label, period_b_tk, fmt,
            )

    # ---------- Sezione SPESA MENSILE (consolidata) ----------
    tk_spend_cur = (tk_cur["spend"] if has_tiktok and tk_cur else 0)
    total_monthly_spent = spend_cur + tk_spend_cur

    ws.cell(row=next_row + 1, column=1, value="═══ SPESA MENSILE ═══").font = Font(bold=True, size=12, color="1F4E78")
    ws.merge_cells(start_row=next_row + 1, start_column=1, end_row=next_row + 1, end_column=4)
    next_row += 3

    # Header
    headers_sm = ["Canale", f"Speso {period_a_label}"]
    for col, h in enumerate(headers_sm, start=1):
        c = ws.cell(row=next_row, column=col, value=h)
        c.font = TABLE_HEADER_FONT; c.fill = TABLE_HEADER_FILL; c.border = BORDER; c.alignment = CENTER
    next_row += 1
    # Rows: Meta, TikTok (se attivo), TOTALE
    rows_spesa = [("Meta", spend_cur)]
    if has_tiktok:
        rows_spesa.append(("TikTok", tk_spend_cur))
    rows_spesa.append(("TOTALE", total_monthly_spent))
    for label, val in rows_spesa:
        is_total = (label == "TOTALE")
        c1 = ws.cell(row=next_row, column=1, value=label)
        c1.font = Font(bold=True) if is_total else TABLE_HEADER_FONT
        c1.border = BORDER; c1.alignment = LEFT
        c2 = ws.cell(row=next_row, column=2, value=fmt_eur(val))
        c2.font = Font(bold=True) if is_total else Font()
        c2.border = BORDER; c2.alignment = CENTER
        next_row += 1
    next_row += 1

    # ---------- Sezione RIEPILOGO SPESA YTD ----------
    budget_annuo = client["budget_annuo"]
    ytd_root = data.get("ytd_spend", {})
    ytd_months = ytd_root.get("months", list(range(1, month + 1)))
    client_ytd = ytd_root.get("by_client", {}).get(client["nome"], {})

    ws.cell(row=next_row + 1, column=1, value="═══ RIEPILOGO SPESA YTD ═══").font = Font(bold=True, size=12, color="1F4E78")
    ws.merge_cells(start_row=next_row + 1, start_column=1, end_row=next_row + 1, end_column=4)
    next_row += 3

    # Header
    for col, h in enumerate(["Mese", "Meta", "TikTok", "Totale"], start=1):
        c = ws.cell(row=next_row, column=col, value=h)
        c.font = TABLE_HEADER_FONT; c.fill = TABLE_HEADER_FILL; c.border = BORDER; c.alignment = CENTER
    next_row += 1

    ytd_meta_tot = 0
    ytd_tiktok_tot = 0
    for m in ytd_months:
        md = client_ytd.get(str(m), {"meta": 0, "tiktok": 0})
        meta_m = md.get("meta", 0) or 0
        tk_m = md.get("tiktok", 0) or 0
        tot_m = meta_m + tk_m
        ytd_meta_tot += meta_m
        ytd_tiktok_tot += tk_m
        c1 = ws.cell(row=next_row, column=1, value=MONTH_IT[m]); c1.font = TABLE_HEADER_FONT; c1.border = BORDER; c1.alignment = LEFT
        c2 = ws.cell(row=next_row, column=2, value=fmt_eur(meta_m)); c2.border = BORDER; c2.alignment = CENTER
        c3 = ws.cell(row=next_row, column=3, value=(fmt_eur(tk_m) if has_tiktok else "—")); c3.border = BORDER; c3.alignment = CENTER
        c4 = ws.cell(row=next_row, column=4, value=fmt_eur(tot_m)); c4.border = BORDER; c4.alignment = CENTER
        next_row += 1

    ytd_tot = ytd_meta_tot + ytd_tiktok_tot
    # Riga TOTALE YTD
    c1 = ws.cell(row=next_row, column=1, value="TOTALE YTD"); c1.font = Font(bold=True); c1.border = BORDER; c1.alignment = LEFT
    c2 = ws.cell(row=next_row, column=2, value=fmt_eur(ytd_meta_tot)); c2.font = Font(bold=True); c2.border = BORDER; c2.alignment = CENTER
    c3 = ws.cell(row=next_row, column=3, value=(fmt_eur(ytd_tiktok_tot) if has_tiktok else "—")); c3.font = Font(bold=True); c3.border = BORDER; c3.alignment = CENTER
    c4 = ws.cell(row=next_row, column=4, value=fmt_eur(ytd_tot)); c4.font = Font(bold=True); c4.border = BORDER; c4.alignment = CENTER
    next_row += 2

    # ---------- Sezione BUDGET TRACKING ANNUO ----------
    cumulative_weight = sum(BUDGET_WEIGHTS[m] for m in ytd_months)
    atteso_ytd = budget_annuo * cumulative_weight / 100
    scarto = ytd_tot - atteso_ytd
    rimanente = budget_annuo - ytd_tot
    tolerance = atteso_ytd * 0.10  # ±10%

    if atteso_ytd == 0:
        status_label, status_color = "— (piano non ancora partito)", "595959"
    elif abs(scarto) <= tolerance:
        status_label, status_color = "✓ IN LINEA col piano", "548235"
    elif scarto < 0:
        status_label, status_color = f"⚠ UNDER SPENDING di {fmt_eur(abs(scarto))}", "BF8F00"
    else:
        status_label, status_color = f"⚠ OVER SPENDING di {fmt_eur(abs(scarto))}", "C00000"

    ws.cell(row=next_row, column=1, value="═══ BUDGET TRACKING ANNUO ═══").font = Font(bold=True, size=12, color="1F4E78")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
    next_row += 2

    tracking_rows = [
        ("Budget annuo", fmt_eur(budget_annuo)),
        (f"Peso cumulato piano (Gen–{MONTH_IT[month]})", f"{cumulative_weight}%"),
        ("Atteso YTD", fmt_eur(atteso_ytd)),
        ("Speso YTD", fmt_eur(ytd_tot)),
        ("Scarto vs piano", f"{'+' if scarto >= 0 else '−'}{fmt_eur(abs(scarto))}"),
        ("Budget rimanente anno", fmt_eur(rimanente)),
    ]
    for label, val in tracking_rows:
        c1 = ws.cell(row=next_row, column=1, value=label); c1.font = TABLE_HEADER_FONT; c1.border = BORDER; c1.alignment = LEFT
        c2 = ws.cell(row=next_row, column=2, value=val); c2.border = BORDER; c2.alignment = CENTER
        ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=4)
        next_row += 1

    # Status pacing (riga evidenziata)
    c1 = ws.cell(row=next_row, column=1, value="Status pacing"); c1.font = Font(bold=True); c1.border = BORDER; c1.alignment = LEFT
    c2 = ws.cell(row=next_row, column=2, value=status_label); c2.font = Font(bold=True, color=status_color, size=11); c2.border = BORDER; c2.alignment = CENTER
    ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=4)
    next_row += 2

    # Variabili per il rational e compatibilità
    total_spent = ytd_tot
    total_expected = atteso_ytd

    # ---------- Sezione PROPOSTA INVESTIMENTO MESE SUCCESSIVO ----------
    next_month = (month % 12) + 1
    next_year = year + 1 if month == 12 else year
    next_weight_pct = BUDGET_WEIGHTS[next_month]
    base_next_month = budget_annuo * next_weight_pct / 100

    if has_tiktok:
        meta_next = base_next_month * META_SHARE
        tiktok_next_base = base_next_month * TIKTOK_SHARE
        tiktok_next = max(TIKTOK_MIN_MONTHLY, tiktok_next_base)
        tk_note = "Min €600/mese" if tiktok_next == TIKTOK_MIN_MONTHLY and tiktok_next_base < TIKTOK_MIN_MONTHLY else f"Split {int(TIKTOK_SHARE*100)}%"
    else:
        meta_next = base_next_month
        tiktok_next = 0
        tk_note = None
    total_next = meta_next + tiktok_next

    ws.cell(row=next_row, column=1, value=f"═══ PROPOSTA INVESTIMENTO {MONTH_IT[next_month].upper()} {next_year} ═══").font = Font(bold=True, size=12, color="1F4E78")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
    next_row += 2

    # Header
    headers_prop = ["Canale", "Investimento Suggerito", "Note"]
    for col, h in enumerate(headers_prop, start=1):
        c = ws.cell(row=next_row, column=col, value=h)
        c.font = TABLE_HEADER_FONT; c.fill = TABLE_HEADER_FILL; c.border = BORDER; c.alignment = CENTER
    # Colonna 4 non usata per questa tabella, ma teniamo il bordo a 3 colonne
    next_row += 1

    # Riga Meta
    c1 = ws.cell(row=next_row, column=1, value="Meta"); c1.font = TABLE_HEADER_FONT; c1.border = BORDER; c1.alignment = LEFT
    c2 = ws.cell(row=next_row, column=2, value=fmt_eur(meta_next)); c2.border = BORDER; c2.alignment = CENTER
    c3 = ws.cell(row=next_row, column=3, value=f"Split {int(META_SHARE*100)}%" if has_tiktok else "100% budget mensile")
    c3.border = BORDER; c3.alignment = CENTER; c3.font = Font(italic=True, size=9, color="595959")
    next_row += 1

    # Riga TikTok (se attivo)
    if has_tiktok:
        c1 = ws.cell(row=next_row, column=1, value="TikTok"); c1.font = TABLE_HEADER_FONT; c1.border = BORDER; c1.alignment = LEFT
        c2 = ws.cell(row=next_row, column=2, value=fmt_eur(tiktok_next)); c2.border = BORDER; c2.alignment = CENTER
        c3 = ws.cell(row=next_row, column=3, value=tk_note); c3.border = BORDER; c3.alignment = CENTER
        c3.font = Font(italic=True, size=9, color="595959")
        next_row += 1

    # Riga TOTALE
    c1 = ws.cell(row=next_row, column=1, value="TOTALE"); c1.font = Font(bold=True); c1.border = BORDER; c1.alignment = LEFT
    c2 = ws.cell(row=next_row, column=2, value=fmt_eur(total_next)); c2.font = Font(bold=True); c2.border = BORDER; c2.alignment = CENTER
    c3 = ws.cell(row=next_row, column=3, value=f"Peso piano {next_weight_pct}%"); c3.border = BORDER; c3.alignment = CENTER
    c3.font = Font(italic=True, size=9, color="595959")
    next_row += 2

    # Nota esplicativa
    nota_txt = (
        f"Calcolo: budget annuo {fmt_eur(budget_annuo)} × peso {MONTH_IT[next_month]} ({next_weight_pct}% del piano AGHC)"
        + (f". Split 80% Meta / 20% TikTok con soglia minima €{int(TIKTOK_MIN_MONTHLY)}/mese su TikTok." if has_tiktok else ".")
    )
    c = ws.cell(row=next_row, column=1, value=nota_txt)
    c.font = Font(italic=True, size=9, color="595959")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
    next_row += 2

    # Rational
    ws.cell(row=next_row, column=1, value="═══ RATIONAL ═══").font = Font(bold=True, size=12, color="1F4E78")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
    next_row += 1
    rational_text = build_rational(
        client["nome"], meta_cur, meta_prev, tk_cur, tk_prev,
        period_a_label, period_b_label, client["confronto_meta"],
        total_spent, total_expected, month, has_tiktok,
    )
    ws.cell(row=next_row, column=1, value=rational_text)
    ws.cell(row=next_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=next_row, column=1).font = Font(size=10)
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row + 8, end_column=4)
    ws.row_dimensions[next_row].height = 140

    # Larghezza colonne
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18


def build_budget_plan_sheet(wb, clients, data, year, month):
    """Tab iniziale con ricalibro budget residuo per i mesi restanti dell'anno.

    Per ogni cliente:
    - budget annuo, speso YTD, residuo disponibile
    - tabella mesi residui con peso piano originale, peso ricalibrato (normalizzato sui
      mesi residui), totale mese suggerito, split Meta/TikTok (80/20 con min €600 TikTok)
    """
    ws = wb.create_sheet(title="Piano Budget Residuo", index=0)

    remaining_months = [m for m in range(month + 1, 13)]
    if not remaining_months:
        ws.cell(row=1, column=1, value="Anno concluso — nessun mese residuo da ricalibrare.").font = Font(italic=True)
        return
    total_remaining_weight = sum(BUDGET_WEIGHTS[m] for m in remaining_months)
    first_next_m = remaining_months[0]

    # Titolo generale
    ws.cell(row=1, column=1, value=f"PIANO BUDGET RESIDUO — da {MONTH_IT[first_next_m]} a Dicembre {year}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value=(
        f"Ricalibrazione dei budget residui sui pesi mensili AGHC — i pesi originali "
        f"({'/'.join(str(BUDGET_WEIGHTS[m]) + '%' for m in remaining_months)}) vengono rinormalizzati "
        f"sul totale residuo {total_remaining_weight}% così che ogni mese riceva una quota proporzionale. "
        f"Split 80% Meta / 20% TikTok dove entrambi attivi, con soglia minima €{int(TIKTOK_MIN_MONTHLY)}/mese TikTok."
    )).font = Font(italic=True, color="595959", size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.row_dimensions[2].height = 32

    ytd_by_client = data.get("ytd_spend", {}).get("by_client", {})
    ytd_months_keys = data.get("ytd_spend", {}).get("months", list(range(1, month + 1)))

    row = 4
    for client in clients:
        has_tiktok = bool(client["tiktok_account_id"])
        budget_annuo = client["budget_annuo"]
        # Calcolo speso YTD
        client_ytd = ytd_by_client.get(client["nome"], {})
        ytd_tot = sum(
            (client_ytd.get(str(m), {}).get("meta", 0) or 0)
            + (client_ytd.get(str(m), {}).get("tiktok", 0) or 0)
            for m in ytd_months_keys
        )
        residuo = budget_annuo - ytd_tot

        # Intestazione cliente
        ws.cell(row=row, column=1, value=client["nome"]).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws.cell(row=row, column=1).alignment = CENTER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

        # Riga info budget cliente
        ws.cell(row=row, column=1, value=(
            f"Budget annuo: {fmt_eur(budget_annuo)}  ·  Speso YTD: {fmt_eur(ytd_tot)}  ·  Residuo disponibile: {fmt_eur(residuo)}"
            + (f"  ·  TikTok: attivo (min €{int(TIKTOK_MIN_MONTHLY)}/mese)" if has_tiktok else "  ·  TikTok: non gestito")
        )).font = Font(italic=True, color="1F4E78", size=10, bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

        # Logica di ricalibro (computata qui per poter stampare l'eventuale warning SOPRA la tabella)
        num_remaining = len(remaining_months)
        warning_line = None
        if has_tiktok:
            tk_fixed_total = TIKTOK_MIN_MONTHLY * num_remaining
            meta_residuo = residuo - tk_fixed_total
            if meta_residuo < 0:
                warning_line = (f"⚠ Residuo {fmt_eur(residuo)} < min TikTok totale ({fmt_eur(tk_fixed_total)}). "
                                f"TikTok manterrà €{int(TIKTOK_MIN_MONTHLY)}/mese fissi, Meta = €0 "
                                f"(sforamento previsto {fmt_eur(-meta_residuo)} da coprire).")
                meta_residuo_eff = 0
            else:
                meta_residuo_eff = meta_residuo
        else:
            meta_residuo_eff = residuo
            tk_fixed_total = 0

        # Warning (se presente) sopra la tabella
        if warning_line:
            ws.cell(row=row, column=1, value=warning_line).font = Font(bold=True, color="C00000", size=10)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            ws.row_dimensions[row].height = 28
            row += 1

        # Header tabella mesi
        headers = ["Mese", "Peso piano", "Peso ricalibrato", "Totale mese", "Meta", "TikTok", "Note"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = TABLE_HEADER_FONT; c.fill = TABLE_HEADER_FILL; c.border = BORDER; c.alignment = CENTER
        row += 1

        # Ricalibro: TikTok fisso €600/mese; Meta = (residuo - TikTok totale) distribuito per peso
        sum_tot = 0.0
        sum_meta = 0.0
        sum_tk = 0.0
        for m in remaining_months:
            peso_orig = BUDGET_WEIGHTS[m]
            peso_ricalibrato = peso_orig / total_remaining_weight * 100 if total_remaining_weight else 0

            if has_tiktok:
                meta_m = meta_residuo_eff * peso_orig / total_remaining_weight if meta_residuo_eff > 0 else 0
                tk_m = TIKTOK_MIN_MONTHLY
                totale_mese = meta_m + tk_m
                note_cell = "TikTok fisso €600 · Meta pro-peso"
            else:
                meta_m = residuo * peso_orig / total_remaining_weight if residuo > 0 else 0
                tk_m = 0
                totale_mese = meta_m
                note_cell = "100% Meta"

            sum_tot += totale_mese
            sum_meta += meta_m
            sum_tk += tk_m

            c1 = ws.cell(row=row, column=1, value=MONTH_IT[m]); c1.font = TABLE_HEADER_FONT; c1.border = BORDER; c1.alignment = LEFT
            c2 = ws.cell(row=row, column=2, value=f"{peso_orig}%"); c2.border = BORDER; c2.alignment = CENTER
            c3 = ws.cell(row=row, column=3, value=f"{peso_ricalibrato:.1f}%"); c3.border = BORDER; c3.alignment = CENTER
            c4 = ws.cell(row=row, column=4, value=fmt_eur(totale_mese)); c4.border = BORDER; c4.alignment = CENTER
            c5 = ws.cell(row=row, column=5, value=fmt_eur(meta_m)); c5.border = BORDER; c5.alignment = CENTER
            c6 = ws.cell(row=row, column=6, value=(fmt_eur(tk_m) if has_tiktok else "—")); c6.border = BORDER; c6.alignment = CENTER
            c7 = ws.cell(row=row, column=7, value=note_cell); c7.border = BORDER; c7.alignment = CENTER
            c7.font = Font(italic=True, size=9, color="595959")
            row += 1

        # Riga TOTALE
        c1 = ws.cell(row=row, column=1, value="TOTALE"); c1.font = Font(bold=True); c1.border = BORDER; c1.alignment = LEFT
        c2 = ws.cell(row=row, column=2, value=f"{sum(BUDGET_WEIGHTS[m] for m in remaining_months)}%"); c2.font = Font(bold=True); c2.border = BORDER; c2.alignment = CENTER
        c3 = ws.cell(row=row, column=3, value="100.0%"); c3.font = Font(bold=True); c3.border = BORDER; c3.alignment = CENTER
        c4 = ws.cell(row=row, column=4, value=fmt_eur(sum_tot)); c4.font = Font(bold=True); c4.border = BORDER; c4.alignment = CENTER
        c5 = ws.cell(row=row, column=5, value=fmt_eur(sum_meta)); c5.font = Font(bold=True); c5.border = BORDER; c5.alignment = CENTER
        c6 = ws.cell(row=row, column=6, value=(fmt_eur(sum_tk) if has_tiktok else "—")); c6.font = Font(bold=True); c6.border = BORDER; c6.alignment = CENTER
        c7 = ws.cell(row=row, column=7, value=""); c7.border = BORDER
        row += 2  # spazio tra clienti

    # Larghezza colonne
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18


def build_index_sheet(wb, clients, year, month):
    """Prima scheda: indice con link rapido ai clienti + overview."""
    ws = wb.create_sheet(title="Indice", index=0)
    ws.cell(row=1, column=1, value=f"AGHC — Report KPI {MONTH_IT[month]} {year}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="1F4E78")
    ws.cell(row=2, column=1, value=f"Generato automaticamente via Windsor.ai (Meta + TikTok)").font = Font(italic=True, color="595959", size=10)

    headers = ["#", "Cliente", "Confronto Meta", "Confronto TikTok", "Budget annuo", "Note"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = TABLE_HEADER_FONT; c.fill = TABLE_HEADER_FILL; c.border = BORDER; c.alignment = CENTER

    for i, client in enumerate(clients):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1).alignment = CENTER
        ws.cell(row=r, column=2, value=client["nome"])
        # Link interno al tab
        ws.cell(row=r, column=2).hyperlink = f"#'{client['nome'][:31]}'!A1"
        ws.cell(row=r, column=2).font = Font(color="1F4E78", underline="single")
        ws.cell(row=r, column=3, value=client["confronto_meta"]).alignment = CENTER
        ws.cell(row=r, column=4, value=client["confronto_tiktok"] or "—").alignment = CENTER
        ws.cell(row=r, column=5, value=fmt_eur(client["budget_annuo"])).alignment = CENTER
        ws.cell(row=r, column=6, value=client["note"]).alignment = LEFT
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 50


def main():
    if len(sys.argv) < 4:
        print("Usage: python generate_sheet.py <data_json> <year> <month> [output.xlsx]")
        sys.exit(1)
    data_path = Path(sys.argv[1])
    year = int(sys.argv[2])
    month = int(sys.argv[3])
    output = Path(sys.argv[4]) if len(sys.argv) > 4 else Path(f"AGHC_Report_KPI_{MONTH_IT[month]}_{year}.xlsx")

    data = json.loads(data_path.read_text())
    wb = Workbook()
    # rimuove il default Sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    build_budget_plan_sheet(wb, CLIENTS, data, year, month)
    for client in CLIENTS:
        build_client_sheet(wb, client, data, year, month)

    wb.save(output)
    print(f"✔ Salvato: {output}")


if __name__ == "__main__":
    main()
