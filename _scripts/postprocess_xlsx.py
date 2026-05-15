#!/usr/bin/env python3
"""postprocess_xlsx.py — Aggiunge note "ⓘ stimata" e "1° mese live" all'xlsx generato.

Replica esattamente il post-processing manuale fatto a maggio 2026 per Aprile 2026:
  1. Per i clienti con reach YoY stimata (reach_estimated_clients nel data JSON), aggiunge
     ⓘ al titolo "ACCOUNT RAGGIUNTI" + commento Excel esplicativo.
  2. Per i clienti con TikTok appena lanciato (cur>0 e prev=0), aggiunge banner
     "📌 TikTok attivato ad <Mese Anno>..." + sostituisce "n/d" con "1° mese live" nella
     sezione TikTok.

Uso:
  python3 postprocess_xlsx.py --xlsx <file.xlsx> --data <data.json> --year <YYYY> --month <MM>
"""
import argparse, json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.comments import Comment

MONTH_IT = {1:"Gennaio",2:"Febbraio",3:"Marzo",4:"Aprile",5:"Maggio",6:"Giugno",
            7:"Luglio",8:"Agosto",9:"Settembre",10:"Ottobre",11:"Novembre",12:"Dicembre"}

# Mappa cliente → tk_account_id (per detection "1° mese live")
TK_BY_CLIENT = {
    "Lunetta": "7498330316248203280",
    "Mare": "7498679494010667009",
    "Marcella Royal": "7499093699838607377",
    "Della Piana": "7504967007843319824",
    "Villa Ermellina": "7612666695502118929",
}

NOTE_FONT_ESTIMATED = Font(italic=True, color="BF8F00", size=9, bold=True)
NOTE_FONT_LAUNCH = Font(italic=True, color="2F5496", size=10, bold=True)
NOTE_FILL_LAUNCH = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")


def add_reach_estimated_note(ws):
    """Aggiunge ⓘ al titolo ACCOUNT RAGGIUNTI con commento Excel."""
    for r in range(1, 30):
        v = ws.cell(row=r, column=1).value
        if v == "ACCOUNT RAGGIUNTI":
            cell = ws.cell(row=r, column=1)
            cell.value = "ACCOUNT RAGGIUNTI  ⓘ"
            cell.comment = Comment(
                "Nota: la reach del periodo precedente (colonna 'Periodo Precedente') è una STIMA, "
                "calcolata applicando il rapporto reach/impressions del periodo corrente "
                "(fallback su MoM quando il corrente non è utilizzabile). Motivo: la Meta "
                "Marketing API non restituisce più il dato reach per periodi oltre 24 mesi.",
                "FMM"
            )
            cell.comment.width = 320
            cell.comment.height = 130
            return True
    return False


def add_tiktok_launch_note(ws, month, year):
    """Aggiunge banner + sostituisce 'n/d' con '1° mese live' nella sezione TikTok."""
    tk_start = None
    for r in range(1, 80):
        v = ws.cell(row=r, column=1).value
        if v and "TIKTOK" in str(v).upper() and "═══" in str(v):
            tk_start = r
            break
    if not tk_start:
        return False
    # Banner sulla riga successiva a "═══ TIKTOK ═══"
    note_row = tk_start + 1
    cell = ws.cell(row=note_row, column=1, value=f"📌 TikTok attivato ad {MONTH_IT[month]} {year} — primo mese live, confronto MoM non disponibile")
    cell.font = NOTE_FONT_LAUNCH
    cell.fill = NOTE_FILL_LAUNCH
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    ws.row_dimensions[note_row].height = 22
    # Sostituisci "n/d" → "1° mese live" nelle righe TikTok della sezione
    for r in range(tk_start, tk_start + 25):
        if ws.cell(row=r, column=1).value == "TikTok":
            cell = ws.cell(row=r, column=4)
            if cell.value == "n/d":
                cell.value = "1° mese live"
                cell.font = Font(italic=True, color="2F5496", size=9, bold=True)
    return True


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True)
    p.add_argument("--data", required=True)
    p.add_argument("--year", type=int, required=True)
    p.add_argument("--month", type=int, required=True)
    args = p.parse_args()

    data = json.loads(Path(args.data).read_text(encoding="utf-8"))
    wb = load_workbook(args.xlsx)

    # 1. Reach stimata
    affected_reach = data.get("reach_estimated_clients", [])
    for nm in affected_reach:
        if nm in wb.sheetnames:
            add_reach_estimated_note(wb[nm])

    # 2. TikTok launch detection
    tk_cur = data.get("tiktok", {}).get("current", {})
    tk_prev = data.get("tiktok", {}).get("prev_mom", {})
    launched = []
    for client_nm, tk_id in TK_BY_CLIENT.items():
        cur = tk_cur.get(tk_id, {})
        prev = tk_prev.get(tk_id, {})
        if (cur.get("spend") or 0) > 0 and (prev.get("spend") or 0) == 0 and (prev.get("impressions") or 0) == 0:
            launched.append(client_nm)

    for nm in launched:
        if nm in wb.sheetnames:
            add_tiktok_launch_note(wb[nm], args.month, args.year)

    wb.save(args.xlsx)
    if affected_reach:
        print(f"✔ ⓘ stimata aggiunta su: {', '.join(affected_reach)}")
    if launched:
        print(f"✔ '1° mese live' applicato a: {', '.join(launched)}")
    if not affected_reach and not launched:
        print("ℹ Nessun post-processing necessario")


if __name__ == "__main__":
    main()
